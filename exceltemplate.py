"""
core/exceltemplate.py

Generates a downloadable .xlsx data-entry template whose columns are built
dynamically from the tags actually found in the uploaded PPTX template
(via core.pptxtagparser.TemplateTagReport).
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from core.pptxtagparser import TemplateTagReport, KNOWN_SINGLE_TAGS


def build_excel_template_bytes(report: TemplateTagReport) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "SlideData"

    columns = []
    for tag in KNOWN_SINGLE_TAGS:
        if tag in report.found_single_tags:
            columns.append(tag)

    for slot in range(1, report.max_speaker_slot + 1):
        columns.append(f"SPEAKER_NAME_{slot}")
        columns.append(f"SPEAKER_TITLE_{slot}")
        columns.append(f"SPEAKER_COMPANY_{slot}")
        columns.append(f"SPEAKER_PHOTO_FILENAME_{slot}")

    columns.append("Require Single Slide")

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[cell.column_letter].width = max(18, len(col_name) + 2)

    ws.cell(row=2, column=columns.index("Require Single Slide") + 1,
            value="Panel 1")
    if "SPEAKER_NAME_1" in columns:
        ws.cell(row=2, column=columns.index("SPEAKER_NAME_1") + 1, value="Jane Doe")
    ws.cell(row=3, column=columns.index("Require Single Slide") + 1,
            value="Panel 1")
    ws.cell(row=4, column=columns.index("Require Single Slide") + 1,
            value="")  # blank = own individual slide

    notes_ws = wb.create_sheet("Instructions")
    notes = [
        "How to fill this sheet:",
        "- Each row = one speaker entry.",
        "- Leave 'Require Single Slide' BLANK if this speaker should get their own slide.",
        "- To combine multiple speakers onto ONE slide, give those rows the SAME label",
        "  in 'Require Single Slide', e.g. 'Panel 1' for all rows that belong on the same slide.",
        "- SESSION_NAME, HALL_NAME, DATE, MAIN_SESSION_DETAILS, SPEAKER_SESSION_DETAILS,",
        "  PLACEHOLDER_1, PLACEHOLDER_2 only need to be filled once per slide/group",
        "  (repeat the same values across grouped rows, or fill only the first row of the group).",
        "- SPEAKER_PHOTO_FILENAME_n should match the filename of a photo you upload separately",
        "  in the app (same auto-matching logic as before).",
    ]
    for i, line in enumerate(notes, start=1):
        notes_ws.cell(row=i, column=1, value=line)
    notes_ws.column_dimensions["A"].width = 100

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
