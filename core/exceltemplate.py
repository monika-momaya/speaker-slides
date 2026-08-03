import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from core.pptxtagparser import TemplateTagReport, KNOWN_SINGLE_TAGS

def build_excel_template_bytes(report: TemplateTagReport) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "SlideData"
    columns = []
    for tag in KNOWN_SINGLE_TAGS:
        if tag in report.found_single_tags:
            columns.append(tag)
    for slot in range(1, report.max_speaker_slot + 1):
        columns += [f"SPEAKER_NAME_{slot}", f"SPEAKER_TITLE_{slot}", f"SPEAKER_COMPANY_{slot}", f"SPEAKER_PHOTO_FILENAME_{slot}"]
    columns.append("Require Single Slide")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[cell.column_letter].width = max(18, len(col_name) + 2)
    notes_ws = wb.create_sheet("Instructions")
    notes = [
        "Each row = one speaker.",
        "Leave Require Single Slide blank for one speaker per slide.",
        "Use the same label, e.g. Panel 1, for rows that belong on the same slide.",
        "Fill session-level fields across grouped rows as needed.",
    ]
    for i, line in enumerate(notes, start=1):
        notes_ws.cell(row=i, column=1, value=line)
    notes_ws.column_dimensions["A"].width = 100
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
