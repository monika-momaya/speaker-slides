from typing import List, Dict
import openpyxl
SINGLE_FIELD_KEYS = ["SESSION_NAME", "HALL_NAME", "DATE", "MAIN_SESSION_DETAILS", "SPEAKER_SESSION_DETAILS", "PLACEHOLDER_1", "PLACEHOLDER_2"]

def read_merge_sheet(file_obj, max_speaker_slot: int):
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb["SlideData"] if "SlideData" in wb.sheetnames else wb.active
    headers = [c.value for c in ws[1]]
    header_index = {h: i for i, h in enumerate(headers) if h}
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or all(v is None or str(v).strip() == '' for v in row):
            continue
        row_dict = {}
        for h, idx in header_index.items():
            row_dict[h] = row[idx] if idx < len(row) else None
        rows.append(row_dict)
    groups: List[Dict] = []
    group_index_by_label: Dict[str, int] = {}
    for row in rows:
        label = str(row.get('Require Single Slide') or '').strip()
        speaker_entry = {
            'name': row.get('SPEAKER_NAME_1') or '',
            'title': row.get('SPEAKER_TITLE_1') or '',
            'company': row.get('SPEAKER_COMPANY_1') or '',
            'photo_key': row.get('SPEAKER_PHOTO_FILENAME_1') or None,
        }
        new_group_data = {k: row.get(k, '') for k in SINGLE_FIELD_KEYS}
        if label:
            if label in group_index_by_label:
                groups[group_index_by_label[label]]['speakers'].append(speaker_entry)
            else:
                new_group_data['speakers'] = [speaker_entry]
                new_group_data['_label'] = label
                groups.append(new_group_data)
                group_index_by_label[label] = len(groups) - 1
        else:
            new_group_data['speakers'] = [speaker_entry]
            new_group_data['_label'] = None
            groups.append(new_group_data)
    overflow_warnings = []
    for g in groups:
        if max_speaker_slot > 0 and len(g['speakers']) > max_speaker_slot:
            label = g.get('_label') or '(individual slide)'
            overflow_warnings.append(f"Group '{label}' has {len(g['speakers'])} speakers but the template only supports {max_speaker_slot} speaker slot(s). Extra speakers will be dropped.")
            g['speakers'] = g['speakers'][:max_speaker_slot]
    return groups, overflow_warnings
